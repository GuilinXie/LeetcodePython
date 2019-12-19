import os
import sys
import time
from datetime import datetime
from openpyxl import Workbook
from decimal import *

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

# create excel object (Workbook) will be save by wb_output.save
wb_output = Workbook()
ws_output = wb_output.active

# configure file names
LogfileNameBase = "start_up_time"

def logfile_init():
    global LogFilexlsx
    for logindex in range(0, 999):
        LogFilexlsx = LogfileNameBase + "_%3.3d.xlsx" % logindex
        if(os.path.exists(LogFilexlsx) == 0):
            break
    print("Writing to Logfile: %s" % LogFilexlsx)

if __name__ == "__main__":
    logfile_init()
    ws_output['A1'] = "Start up time measure"
    LineHeader = 4
    item = ['test1', 'test2', 'test3', 'test4']
    for t in item:
        ws_output['A'+str(LineHeader)] = t
        LineHeader += 1
    MinColumn = get_column_letter(len(item) + 1)
    MaxColumn = get_column_letter(len(item) + 2)
    AvgColumn = get_column_letter(len(item) + 3)
    ws_output[MinColumn + str(LineHeader)] = "Min"
    ws_output[MaxColumn + str(LineHeader)] = "Max"
    ws_output[AvgColumn + str(LineHeader)] = "Avg"